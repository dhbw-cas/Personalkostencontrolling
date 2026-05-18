"""Excel-Datei Writer für Zielformat."""

import logging
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ExcelWriterError(Exception):
    """Fehler beim Schreiben von Excel-Dateien."""


class ExcelWriter:
    """Schreibt transformierte Daten in Excel-Zielformat."""

    def __init__(self, output_path: Path) -> None:
        """Initialisiert den Excel Writer.

        Args:
            output_path: Pfad für die Ziel-Excel-Datei
        """
        self.output_path = output_path
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

    def write_data(self, df: pd.DataFrame, sheet_name: str = "Buchungsdaten") -> None:
        """Schreibt DataFrame in Excel-Datei mit SAP-konformer Formatierung.

        Args:
            df: DataFrame mit transformierten Daten
            sheet_name: Name des Arbeitsblatts

        Raises:
            ExcelWriterError: Bei Fehlern beim Schreiben
        """
        try:
            logger.info(f"Schreibe Excel-Datei: {self.output_path}")
            self._write_workbook({sheet_name: df})
            logger.info(f"Excel-Datei erfolgreich geschrieben: {len(df)} Zeilen")
        except Exception as e:
            raise ExcelWriterError(f"Fehler beim Schreiben der Excel-Datei: {e}") from e

    def _apply_formatting(self, worksheet: Any, df: pd.DataFrame) -> None:
        """Wendet SAP-konforme Formatierung auf das Arbeitsblatt an.

        Args:
            worksheet: openpyxl Worksheet
            df: Original DataFrame für Dimensionen
        """
        # Header-Formatierung
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center")

        # Formatiere Header-Zeile
        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = header_alignment

        # Spaltenbreiten anpassen
        column_widths = self._calculate_column_widths(df)

        for i, width in enumerate(column_widths, 1):
            column_letter = self._get_column_letter(i)
            worksheet.column_dimensions[column_letter].width = width

        # Datenformat für spezielle Spalten
        self._format_data_columns(worksheet, df)

    def _calculate_column_widths(self, df: pd.DataFrame) -> list[float]:
        """Berechnet optimale Spaltenbreiten basierend auf Inhalt.

        Args:
            df: DataFrame für Breitenberechnung

        Returns:
            Liste der Spaltenbreiten
        """
        widths = []

        for column in df.columns:
            # Maximal-Breite: Header vs. längster Wert
            header_length = len(str(column))

            if not df[column].empty:
                max_content_length = df[column].astype(str).str.len().max()
            else:
                max_content_length = 0

            # Nimm das Maximum, aber begrenzt auf sinnvolle Werte
            width = max(header_length, max_content_length, 8)  # Mindestens 8
            width = min(width, 50)  # Maximal 50

            widths.append(width * 1.2)  # Etwas Puffer

        return widths

    def _format_data_columns(self, worksheet: Any, df: pd.DataFrame) -> None:
        """Formatiert spezifische Datenspalten nach SAP-Anforderungen.

        Args:
            worksheet: openpyxl Worksheet
            df: DataFrame mit den Daten
        """
        # Datumsformatierung für Rechnungs- und Buchungsdatum
        date_columns = ["Rechnungsdatum", "Buchungsdatum"]
        number_columns = ["Betrag Hausw", "Steuerbetrag", "Position", "Zeile"]

        for row_idx in range(2, len(df) + 2):  # Start bei Zeile 2 (nach Header)
            for col_idx, column_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)

                # Datumsformatierung
                if column_name in date_columns:
                    cell.number_format = "DD.MM.YYYY"

                # Zahlenformatierung
                elif column_name in number_columns:
                    if column_name in ["Betrag Hausw", "Steuerbetrag"]:
                        numeric_value = self._to_excel_decimal(cell.value)
                        if numeric_value is not None:
                            cell.value = numeric_value
                        cell.number_format = "#,##0.00"  # Währungsformat
                    else:
                        numeric_value = self._to_excel_integer(cell.value)
                        if numeric_value is not None:
                            cell.value = numeric_value
                        cell.number_format = "0"  # Ganzzahl

                # Text rechtsbündig für bestimmte Felder
                if column_name in ["Belegnummer", "Buchungskreis", "Kreditor"]:
                    cell.alignment = Alignment(horizontal="right")

    def _to_excel_decimal(self, value: Any) -> float | None:
        """Konvertiert Zellenwerte in echte numerische Excel-Dezimalwerte."""
        decimal_value = self._parse_decimal(value)
        if decimal_value is None:
            return None
        return float(decimal_value)

    def _to_excel_integer(self, value: Any) -> int | None:
        """Konvertiert Zellenwerte in echte numerische Excel-Ganzzahlen."""
        decimal_value = self._parse_decimal(value)
        if decimal_value is None:
            return None
        return int(decimal_value)

    def _parse_decimal(self, value: Any) -> Decimal | None:
        """Parst Zahlwerte robust aus Text- und Zahlenzellen."""
        if value is None:
            return None

        if isinstance(value, str):
            normalized = value.strip()
            if not normalized:
                return None
            normalized = normalized.replace(" ", "")
            normalized = normalized.replace("€", "").replace("EUR", "")
            if "," in normalized and "." in normalized:
                if normalized.rfind(",") > normalized.rfind("."):
                    normalized = normalized.replace(".", "").replace(",", ".")
                else:
                    normalized = normalized.replace(",", "")
            else:
                normalized = normalized.replace(",", ".")
        else:
            normalized = str(value)

        try:
            return Decimal(normalized)
        except (InvalidOperation, ValueError, TypeError):
            logger.warning("Konnte Excel-Zahl '%s' nicht numerisch schreiben", value)
            return None

    def _get_column_letter(self, col_num: int) -> str:
        """Konvertiert Spaltennummer zu Excel-Buchstabe.

        Args:
            col_num: Spaltennummer (1-basiert)

        Returns:
            Excel-Spaltenbezeichnung (A, B, C, ...)
        """
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(65 + (col_num % 26)) + result
            col_num //= 26
        return result

    def write_with_template_check(
        self,
        df: pd.DataFrame,
        expected_columns: list[str],
        sheet_name: str = "Buchungsdaten",
        extra_sheets: dict[str, pd.DataFrame] | None = None,
    ) -> None:
        """Schreibt Daten und validiert gegen erwartete Spaltenstruktur.

        Args:
            df: DataFrame mit transformierten Daten
            expected_columns: Liste der erwarteten Spaltennamen in Reihenfolge
            sheet_name: Name des Arbeitsblatts

        Raises:
            ExcelWriterError: Bei fehlenden oder falschen Spalten
        """
        # Validierung der Spaltenstruktur
        sheets: dict[str, pd.DataFrame] = {}
        ordered_main = self._prepare_dataframe(df, expected_columns)
        sheets[sheet_name] = ordered_main

        if extra_sheets:
            for name, extra_df in extra_sheets.items():
                ordered_extra = self._prepare_dataframe(extra_df, expected_columns)
                sheets[name] = ordered_extra

        self._write_workbook(sheets)

        logger.info(
            f"Template-validierte Excel-Datei geschrieben: {len(ordered_main)} Zeilen"
        )

    def _prepare_dataframe(
        self, df: pd.DataFrame, expected_columns: list[str]
    ) -> pd.DataFrame:
        """Reindexed DataFrame according to expected columns with validation."""
        missing_columns = set(expected_columns) - set(df.columns)
        if missing_columns:
            raise ExcelWriterError(f"Fehlende Spalten: {missing_columns}")

        extra_columns = set(df.columns) - set(expected_columns)
        if extra_columns:
            logger.warning(f"Zusätzliche Spalten werden ignoriert: {extra_columns}")

        ordered_df = df.reindex(columns=expected_columns)
        if ordered_df.empty:
            # Stelle sicher, dass zumindest Header vorhanden sind
            return pd.DataFrame(columns=expected_columns)
        return ordered_df

    def _write_workbook(self, sheets: dict[str, pd.DataFrame]) -> None:
        """Schreibt mehrere Arbeitsblätter in eine Excel-Datei."""
        workbook = Workbook()
        first_sheet = True

        for name, df in sheets.items():
            if first_sheet:
                worksheet = workbook.active
                worksheet.title = name
                first_sheet = False
            else:
                worksheet = workbook.create_sheet(title=name)

            for row in dataframe_to_rows(df, index=False, header=True):
                worksheet.append(row)

            self._apply_formatting(worksheet, df)

        workbook.save(self.output_path)


def create_sap_template_columns() -> list[str]:
    """Erstellt die Standard-Spaltenreihenfolge für SAP LBV Import.

    Returns:
        Liste der Spaltennamen in der korrekten Reihenfolge
    """
    return [
        "Belegnummer",
        "Geschäftsjahr",
        "Zeileart",
        "Buchungskreis",
        "Belegart",
        "Position",
        "Rechnungsdatum",
        "Buchungsdatum",
        "Buchungsperiode",
        "Referenz",
        "Belegkopftext",
        "Debitor",
        "Kreditor",
        "Text",
        "Zuordnung",
        "Zahlweg",
        "Zahlungsbed",
        "Mahnbereich",
        "GeschBereich",
        "Steuerkennz.",
        "Soll/Haben",
        "Betrag Hausw",
        "Steuerbetrag",
        "Hauptbuch",
        "Kostenstelle",
        "Auftrag",
        "PSP-Element",
        "Fonds",
        "Währung",
        "Referenzschl 1",
        "GrpId",
        "Status",
        "Icon",
        "Ergebnis",
        "Zeile",
        "Fehler",
        "Warnungen",
        "Informationen",
    ]
