"""Excel-Datei Einleser für Quelldaten."""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelReaderError(Exception):
    """Fehler beim Einlesen von Excel-Dateien."""


class ExcelReader:
    """Liest Excel-Quelldateien ein und stellt die Daten strukturiert bereit."""

    def __init__(self, file_path: Path) -> None:
        """Initialisiert den Excel Reader.

        Args:
            file_path: Pfad zur Excel-Quelldatei

        Raises:
            ExcelReaderError: Wenn die Datei nicht existiert oder nicht lesbar ist
        """
        self.file_path = file_path

        if not self.file_path.exists():
            raise ExcelReaderError(f"Datei nicht gefunden: {file_path}")

        if self.file_path.suffix.lower() not in [".xlsx", ".xls"]:
            raise ExcelReaderError(f"Ungültiges Dateiformat: {file_path.suffix}")

    def read_data(self, sheet_name: str | int = 0) -> pd.DataFrame:
        """Liest die Quelldaten aus der Excel-Datei.

        Args:
            sheet_name: Name oder Index des Arbeitsblatts (default: erstes Blatt)

        Returns:
            DataFrame mit den Quelldaten

        Raises:
            ExcelReaderError: Bei Fehlern beim Einlesen
        """
        try:
            logger.info(f"Lade Excel-Datei: {self.file_path}")

            # Versuche zuerst mit openpyxl für bessere Formatierung
            try:
                workbook = load_workbook(self.file_path, data_only=True)

                # Wähle Arbeitsblatt
                if isinstance(sheet_name, str):
                    if sheet_name not in workbook.sheetnames:
                        raise ExcelReaderError(
                            f"Arbeitsblatt '{sheet_name}' nicht gefunden"
                        )
                    worksheet = workbook[sheet_name]
                else:
                    worksheet = workbook.worksheets[sheet_name]

                # Konvertiere zu DataFrame
                data = []
                headers = []

                # Erste Zeile als Header
                first_row = next(worksheet.iter_rows(values_only=True))
                headers = [
                    str(cell) if cell is not None else f"Col_{i}"
                    for i, cell in enumerate(first_row)
                ]

                # Datenzeilen
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    data.append(list(row))

                df = pd.DataFrame(data, columns=headers)

            except Exception as e:
                logger.warning(f"openpyxl fehlgeschlagen, versuche pandas: {e}")
                # Fallback zu pandas
                df = pd.read_excel(self.file_path, sheet_name=sheet_name)

            # Validierung
            if df.empty:
                raise ExcelReaderError("Excel-Datei ist leer")

            logger.info(
                f"Erfolgreich geladen: {len(df)} Zeilen, {len(df.columns)} Spalten"
            )

            # Spaltennamen zu Standard-Excel-Notation (A, B, C, ...)
            df_with_excel_columns = self._add_excel_column_mapping(df)

            return df_with_excel_columns

        except Exception as e:
            if isinstance(e, ExcelReaderError):
                raise
            raise ExcelReaderError(f"Fehler beim Einlesen der Excel-Datei: {e}")

    def _add_excel_column_mapping(self, df: pd.DataFrame) -> pd.DataFrame:
        """Fügt Excel-Spaltenbezeichnungen (A, B, C...) als zusätzliche Referenz hinzu.

        Args:
            df: Original DataFrame

        Returns:
            DataFrame mit zusätzlichen Excel-Spalten-Referenzen
        """
        # Erstelle eine Kopie
        result_df = df.copy()

        # Füge Excel-Spaltenbezeichnungen als Index-Attribut hinzu
        excel_columns = {}
        for i in range(len(df.columns)):
            excel_col = self._number_to_excel_column(i + 1)
            excel_columns[excel_col] = df.columns[i]

        # Speichere Mapping als Attribut des DataFrames
        result_df.attrs["excel_column_mapping"] = excel_columns

        logger.debug(f"Excel-Spalten-Mapping: {excel_columns}")

        return result_df

    def _number_to_excel_column(self, n: int) -> str:
        """Konvertiert eine Spaltennummer zu Excel-Spaltenbezeichnung (1->A, 2->B, etc.).

        Args:
            n: Spaltennummer (1-basiert)

        Returns:
            Excel-Spaltenbezeichnung (A, B, C, ..., Z, AA, AB, ...)
        """
        result = ""
        while n > 0:
            n -= 1  # Mache 0-basiert
            result = chr(65 + (n % 26)) + result
            n //= 26
        return result

    def get_column_by_excel_name(
        self, df: pd.DataFrame, excel_column: str
    ) -> pd.Series:
        """Gibt eine Spalte anhand der Excel-Bezeichnung zurück.

        Args:
            df: DataFrame mit Excel-Spalten-Mapping
            excel_column: Excel-Spaltenbezeichnung (A, B, C, ...)

        Returns:
            Pandas Series der entsprechenden Spalte

        Raises:
            ExcelReaderError: Wenn die Spalte nicht existiert
        """
        if "excel_column_mapping" not in df.attrs:
            raise ExcelReaderError("DataFrame hat kein Excel-Spalten-Mapping")

        mapping = df.attrs["excel_column_mapping"]

        if excel_column not in mapping:
            available_cols = list(mapping.keys())
            raise ExcelReaderError(
                f"Excel-Spalte '{excel_column}' nicht gefunden. "
                f"Verfügbare Spalten: {available_cols}"
            )

        original_column_name = mapping[excel_column]
        return df[original_column_name]

    def validate_required_columns(
        self, df: pd.DataFrame, required_excel_columns: list[str]
    ) -> None:
        """Validiert, dass alle benötigten Excel-Spalten vorhanden sind.

        Args:
            df: DataFrame zum Validieren
            required_excel_columns: Liste der benötigten Excel-Spalten (z.B. ["A", "B", "C"])

        Raises:
            ExcelReaderError: Wenn benötigte Spalten fehlen
        """
        if "excel_column_mapping" not in df.attrs:
            raise ExcelReaderError("DataFrame hat kein Excel-Spalten-Mapping")

        mapping = df.attrs["excel_column_mapping"]
        available_columns = set(mapping.keys())
        required_columns = set(required_excel_columns)

        missing_columns = required_columns - available_columns

        if missing_columns:
            raise ExcelReaderError(
                f"Benötigte Spalten fehlen: {sorted(missing_columns)}. "
                f"Verfügbare Spalten: {sorted(available_columns)}"
            )
